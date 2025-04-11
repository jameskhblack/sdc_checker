"""Applies styling to the generated Excel sheets using openpyxl."""

import logging
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from typing import Tuple

# Assuming ConfigModel and exceptions are importable from sibling modules
from .config import ConfigModel
from .exceptions import StylingError

# Get the logger instance configured in utils
logger = logging.getLogger("sdc_checker.styler")

# --- Styling Constants ---
TITLE_FONT = Font(name='Calibri', size=13, bold=True)
HEADER_FONT = Font(name='Calibri', size=11, bold=True)
DEFAULT_FONT = Font(name='Calibri', size=11)
AGGREGATED_FONT = Font(name='Calibri', size=11, bold=True)  # Bold font for Aggregated cells
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=False)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=False) # Numbers usually right-aligned

THIN_BORDER_SIDE = Side(border_style="thin", color="000000")
# Individual borders for combining
BOTTOM_BORDER = Border(bottom=THIN_BORDER_SIDE)
RIGHT_BORDER = Border(right=THIN_BORDER_SIDE)
BOTTOM_RIGHT_BORDER = Border(bottom=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE) # Pre-combined for convenience if needed
NO_BORDER = Border()

# Fill for SDC failures (e.g., light yellow)
FAIL_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Number formats
FORMAT_INTEGER = numbers.FORMAT_NUMBER # '#,##0'
FORMAT_PERCENTAGE = numbers.FORMAT_PERCENTAGE_00 # '0.00%'
FORMAT_NUMBER_1DP = '#,##0.0' # Explicit format string for one decimal place

# --- Helper Functions for Border Logic ---

def _is_innermost_col_header_row(r_idx: int, col_header_end_row: int) -> bool:
    """Checks if the row index corresponds to the last header row."""
    # The last header row is one *before* the col_header_end_row
    return r_idx == col_header_end_row - 1

def _is_innermost_row_index_col(c_idx: int, data_col_start: int) -> bool:
    """Checks if the column index corresponds to the last index column."""
    return c_idx == data_col_start - 1

def _is_last_row_in_outer_group(df_r_idx: int, row_indices: pd.Index, num_index_cols: int, total_df_rows: int) -> bool:
    """
    Checks if the DataFrame row is the last in its group, based on all index
    levels except the innermost one.
    """
    # This logic only applies if there are multiple index levels (num_index_cols > 1)
    # and we are not looking at an invalid index or the very last row (which is handled separately)
    if num_index_cols <= 1 or df_r_idx < 0 or df_r_idx >= total_df_rows - 1:
        return False

    # If we reached here, df_r_idx is a valid index before the last row
    # Compare levels 0 to n-2 (which is index num_index_cols - 1)
    levels_to_compare = num_index_cols - 1 # This is the number of outer levels

    current_row_index = row_indices[df_r_idx]
    next_row_index = row_indices[df_r_idx + 1]

    # Ensure comparison works for both MultiIndex (tuples) and single Index
    if isinstance(current_row_index, tuple):
        # Compare the tuple slices containing the outer levels
        return current_row_index[:levels_to_compare] != next_row_index[:levels_to_compare]
    else:
        # If not a tuple, it's a single index. The check num_index_cols <= 1 should have caught this.
        # If somehow reached, return False as there are no outer groups.
        return False

def _is_last_col_in_outer_group(df_c_idx: int, col_indices: pd.Index, num_header_rows: int, total_df_cols: int) -> bool:
    """
    Checks if the DataFrame column is the last in its group, based on all header
    levels except the innermost one.
    """
    # This logic only applies if there are multiple header levels (num_header_rows > 1)
    # and we are not looking at an invalid index or the very last column (which is handled separately)
    if num_header_rows <= 1 or df_c_idx < 0 or df_c_idx >= total_df_cols - 1:
        return False

    # If we reached here, df_c_idx is a valid index before the last column
    # Compare levels 0 to n-2 (which is index num_header_rows - 1)
    levels_to_compare = num_header_rows - 1 # This is the number of outer levels

    current_col_index = col_indices[df_c_idx]
    next_col_index = col_indices[df_c_idx + 1]

    # Ensure comparison works for both MultiIndex (tuples) and single Index
    if isinstance(current_col_index, tuple):
        # Compare the tuple slices containing the outer levels
        return current_col_index[:levels_to_compare] != next_col_index[:levels_to_compare]
    else:
        # If not a tuple, it's a single index. The check num_header_rows <= 1 should have caught this.
        # If somehow reached, return False as there are no outer groups.
        return False

def _is_aggregated_value(value):
    """
    Checks if a cell value is 'Aggregated'.
    
    Args:
        value: The cell value to check
        
    Returns:
        bool: True if the value is 'Aggregated', False otherwise
    """
    if isinstance(value, str) and value == 'Aggregated':
        return True
    return False


# --- Main Styling Function ---

def apply_styles(workbook: Workbook,
                 worksheet: Worksheet,
                 original_df: pd.DataFrame, # DF *before* unmerge hack
                 config: ConfigModel,
                 table_type: str) -> None:
    """
    Applies formatting to the specified worksheet based on the table type.

    Args:
        workbook: The openpyxl Workbook object.
        worksheet: The openpyxl Worksheet object to style.
        original_df: The original DataFrame (before unmerge hack) used for structure reference.
        config: The validated configuration object.
        table_type: Type of table ('primary', 'sdc_detailed', 'pass_fail').

    Raises:
        StylingError: If errors occur during styling application.
    """
    logger.info(f"Applying styles to sheet '{worksheet.title}' (type: {table_type})...")

    if original_df.empty:
        logger.warning(f"Original DataFrame for sheet '{worksheet.title}' is empty. Skipping styling.")
        return

    try:
        # --- Determine Offsets and Dimensions ---
        header_rows = original_df.columns.nlevels
        index_cols = original_df.index.nlevels

        title_row = 1
        header_row_start = title_row + 1
        col_header_end_row = header_row_start + header_rows # Row *after* the last header row
        data_row_start = header_row_start + header_rows

        index_col_start = 1
        data_col_start = index_col_start + index_cols

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Get DF structure info needed for border logic
        total_df_rows = len(original_df.index) if original_df.index is not None else 0
        total_df_cols = len(original_df.columns) if original_df.columns is not None else 0
        # Handle cases where index or columns might be None or not MultiIndex safely
        row_indices = original_df.index if isinstance(original_df.index, pd.MultiIndex) or isinstance(original_df.index, pd.Index) else None
        col_indices = original_df.columns if isinstance(original_df.columns, pd.MultiIndex) or isinstance(original_df.columns, pd.Index) else None


        logger.debug(f"Calculated styling parameters:")
        logger.debug(f"  Header rows: {header_rows}, Index cols: {index_cols}")
        logger.debug(f"  Header row start: {header_row_start}, Data row start: {data_row_start}")
        logger.debug(f"  Index col start: {index_col_start}, Data col start: {data_col_start}")
        logger.debug(f"  Max row: {max_row}, Max col: {max_col}")
        logger.debug(f"  Total DF rows: {total_df_rows}, Total DF cols: {total_df_cols}")
        # Adjust data_row_start if index is MultiIndex (pandas adds extra row for index names)
        if index_cols > 1:
            data_row_start += 1
            logger.debug(f"  Adjusted data_row_start for MultiIndex rows: {data_row_start}")


        # --- 1. Style Title ---
        title_cell_coord = "A1"
        try:
            title_cell = worksheet.cell(row=1, column=1)
            if title_cell.value is not None:
                title_cell.font = TITLE_FONT
                logger.debug(f"Applied title font to cell {title_cell_coord}")
            else:
                logger.warning(f"Title cell {title_cell_coord} found but is empty, font not applied.")
        except Exception as e:
             logger.warning(f"Could not access or style title cell {title_cell_coord}: {e}")

        # --- 2. Rename NaN/None Headers to 'Missing' ---
        logger.debug("Renaming NaN/None headers...")
        # Column Headers
        if max_col >= data_col_start: # Ensure there are data columns
            # Use col_header_end_row, which is independent of the data_row_start adjustment
            for r_idx in range(header_row_start, col_header_end_row): # <<< CORRECTED RANGE
                for c_idx in range(data_col_start, max_col + 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    if cell.value is None or cell.value == '' or pd.isna(cell.value):
                        cell.value = 'Missing'
        # Row Index Headers
        if max_row >= data_row_start: # Ensure there are data rows
            for c_idx in range(index_col_start, data_col_start):
                for r_idx in range(data_row_start, max_row + 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    if cell.value is None or cell.value == '' or pd.isna(cell.value):
                        cell.value = 'Missing'


        # --- 3. Apply Default Font and Alignment ---
        logger.debug("Applying default font and alignments...")
        if max_row >= header_row_start and max_col >= index_col_start: # Check if there are any cells to style
            for r_idx in range(header_row_start, max_row + 1):
                 for c_idx in range(index_col_start, max_col + 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.font = DEFAULT_FONT
                    if c_idx < data_col_start: # Index columns
                        cell.alignment = LEFT_ALIGN
                    else: # Header / Data columns
                        if r_idx < data_row_start: # Header rows
                             cell.alignment = LEFT_ALIGN # Keep headers left-aligned
                             cell.font = HEADER_FONT # Bold headers
                        else: # Data cells
                             cell.alignment = RIGHT_ALIGN # Right align data


        # --- 4. Apply Number Formats ---
        logger.debug("Applying number formats...")
        if table_type == 'sdc_detailed' and max_row >= data_row_start and max_col >= data_col_start:
            indicator_col_idx = data_col_start - 1 # Last index col
            for r_idx in range(data_row_start, max_row + 1):
                try: # Add try-except for cell access
                    indicator_cell = worksheet.cell(row=r_idx, column=indicator_col_idx)
                    indicator_name = indicator_cell.value
                    num_format = FORMAT_NUMBER_1DP # Default
                    if indicator_name == 'Count':
                        num_format = FORMAT_INTEGER
                    elif indicator_name in ['Dominance (1,k) check', 'Dominance p% check']:
                        num_format = FORMAT_PERCENTAGE

                    for c_idx in range(data_col_start, max_col + 1):
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        # Apply format only if the cell contains a number
                        if isinstance(cell.value, (int, float, np.number)) and not pd.isna(cell.value):
                            cell.number_format = num_format
                except IndexError:
                     logger.warning(f"IndexError accessing cell at row {r_idx}, col {indicator_col_idx} or beyond during number formatting.")
                     continue # Skip to next row if index is out of bounds

        elif table_type == 'primary' and max_row >= data_row_start and max_col >= data_col_start:
             for r_idx in range(data_row_start, max_row + 1):
                 for c_idx in range(data_col_start, max_col + 1):
                    try: # Add try-except for cell access
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        if isinstance(cell.value, (int, float, np.number)) and not pd.isna(cell.value):
                             cell.number_format = FORMAT_NUMBER_1DP
                    except IndexError:
                         logger.warning(f"IndexError accessing cell at row {r_idx}, col {c_idx} during primary number formatting.")
                         continue # Skip to next cell in row if index is out of bounds
        # No specific number format needed for 'pass_fail' table


        # --- 5. Conditional Formatting (SDC Detailed Table Only) ---
        logger.debug("Applying conditional formatting...")
        if table_type == 'sdc_detailed' and max_row >= data_row_start and max_col >= data_col_start:
            # Ensure config and sdc_rules exist
            if config and hasattr(config, 'sdc_rules') and config.sdc_rules:
                rules = config.sdc_rules
                indicator_col_idx = data_col_start - 1
                data_range_str = f"{get_column_letter(data_col_start)}{data_row_start}:{get_column_letter(max_col)}{max_row}"
                logger.debug(f"Applying CF to range: {data_range_str}")

                # Ensure thresholds are valid numbers before creating rules
                try:
                    primary_thresh = float(rules.primary_threshold)
                    dom1k_thresh = float(rules.dominance_1k)
                    domp_thresh = float(rules.dominance_p_percent)

                    rule_count = CellIsRule(operator='lessThan', formula=[primary_thresh], stopIfTrue=False, fill=FAIL_FILL)
                    rule_dom1k = CellIsRule(operator='greaterThan', formula=[dom1k_thresh], stopIfTrue=False, fill=FAIL_FILL)
                    rule_domp = CellIsRule(operator='lessThan', formula=[domp_thresh], stopIfTrue=False, fill=FAIL_FILL)

                    for r_idx in range(data_row_start, max_row + 1):
                        try: # Add try-except for cell access
                            indicator_name = worksheet.cell(row=r_idx, column=indicator_col_idx).value
                            row_range = f"{get_column_letter(data_col_start)}{r_idx}:{get_column_letter(max_col)}{r_idx}"

                            if indicator_name == 'Count':
                                worksheet.conditional_formatting.add(row_range, rule_count)
                            elif indicator_name == 'Dominance (1,k) check':
                                worksheet.conditional_formatting.add(row_range, rule_dom1k)
                            elif indicator_name == 'Dominance p% check':
                                worksheet.conditional_formatting.add(row_range, rule_domp)
                        except IndexError:
                            logger.warning(f"IndexError accessing cell at row {r_idx}, col {indicator_col_idx} or beyond during conditional formatting.")
                            continue # Skip to next row
                except (ValueError, TypeError) as rule_err:
                     logger.error(f"Invalid threshold value in sdc_rules for conditional formatting: {rule_err}")
                except Exception as cf_err: # Catch other potential errors during CF
                     logger.error(f"Error applying conditional formatting rules: {cf_err}", exc_info=True)
            else:
                 logger.warning("SDC rules not found in config or invalid, skipping conditional formatting.")


        # --- 6. Apply Borders ---
        logger.debug("Applying custom borders...")
        if max_row >= header_row_start and max_col >= index_col_start: # Check if there are any cells to apply borders to
            # Pre-calculate if group checks are needed at all
            check_row_groups = index_cols > 1 and row_indices is not None and total_df_rows > 0
            check_col_groups = header_rows > 1 and col_indices is not None and total_df_cols > 0

            for r_idx in range(header_row_start, max_row + 1):
                for c_idx in range(index_col_start, max_col + 1):
                    apply_bottom = False
                    apply_right = False

                    # Map worksheet cell index to DataFrame index (0-based)
                    # These indices are only valid/meaningful for data cells when checking groups
                    df_r_idx = r_idx - data_row_start
                    df_c_idx = c_idx - data_col_start

                    # 1. Check Group Borders
                    # Check row group border (only applies if multi-index rows exist and it's a data row or beyond)
                    is_last_row = (df_r_idx == total_df_rows - 1)
                    if check_row_groups and r_idx >= data_row_start:
                         if is_last_row or _is_last_row_in_outer_group(df_r_idx, row_indices, index_cols, total_df_rows):
                             apply_bottom = True

                    # Check column group border (only applies if multi-index columns exist and it's a data column or beyond)
                    is_last_col = (df_c_idx == total_df_cols - 1)
                    if check_col_groups and c_idx >= data_col_start:
                         if is_last_col or _is_last_col_in_outer_group(df_c_idx, col_indices, header_rows, total_df_cols):
                             apply_right = True

                    # 2. Check Innermost Heading Borders
                    # Pass the explicit col_header_end_row instead of relying on data_row_start
                    if _is_innermost_col_header_row(r_idx, col_header_end_row): # <<< UPDATED CALL
                        apply_bottom = True
                    if _is_innermost_row_index_col(c_idx, data_col_start):
                        apply_right = True

                    # 3. Apply the calculated border
                    try: # Add try-except for cell access
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        final_border = Border() # Start with no border
                        if apply_bottom:
                            final_border.bottom = THIN_BORDER_SIDE
                        if apply_right:
                            final_border.right = THIN_BORDER_SIDE
                        # Only apply if different from default NO_BORDER to potentially improve performance
                        if final_border != NO_BORDER:
                             cell.border = final_border
                        else:
                             # Ensure explicitly no border if none applied, overriding potential previous formats
                             cell.border = NO_BORDER
                    except IndexError:
                         logger.warning(f"IndexError accessing cell at row {r_idx}, col {c_idx} during border application.")
                         continue # Skip to next cell


        # --- 7. Set Column Widths ---
        logger.debug("Setting column widths...")
        # Index columns
        if max_col >= index_col_start:
            for c_idx in range(index_col_start, min(data_col_start, max_col + 1)): # Ensure loop doesn't exceed max_col
                try:
                    col_letter = get_column_letter(c_idx)
                    is_indicator_col = (table_type == 'sdc_detailed' and c_idx == (data_col_start - 1))
                    width = 21 if is_indicator_col else 10.5
                    worksheet.column_dimensions[col_letter].width = width
                    # logger.debug(f"Set width for index column {col_letter} to {width}") # Reduce verbosity
                except Exception as width_err:
                     logger.warning(f"Error setting width for index column {c_idx}: {width_err}")

        # Data columns
        if max_col >= data_col_start:
            for c_idx in range(data_col_start, max_col + 1):
                try:
                    col_letter = get_column_letter(c_idx)
                    worksheet.column_dimensions[col_letter].width = 10.5
                    # logger.debug(f"Set width for data column {col_letter} to {10.5}") # Reduce verbosity
                except Exception as width_err:
                     logger.warning(f"Error setting width for data column {c_idx}: {width_err}")


        # --- Apply Aggregated Font to Aggregated Rows and Columns ---
        logger.debug("Applying bold font to 'Aggregated' rows and columns...")
        if max_row >= data_row_start and max_col >= data_col_start:
            # First, collect all the aggregated row and column indices
            aggregated_row_indices = []
            aggregated_col_indices = []
            
            # Find all aggregated rows
            for r_idx in range(data_row_start, max_row + 1):
                is_aggregated_row = False
                # Check index columns for 'Aggregated' value
                for c_idx in range(index_col_start, data_col_start):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    if _is_aggregated_value(cell.value):
                        is_aggregated_row = True
                        break
                if is_aggregated_row:
                    aggregated_row_indices.append(r_idx)
            
            # Find all aggregated columns
            for c_idx in range(data_col_start, max_col + 1):
                is_aggregated_col = False
                # Check header rows for 'Aggregated' value
                for r_idx in range(header_row_start, data_row_start):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    if _is_aggregated_value(cell.value):
                        is_aggregated_col = True
                        break
                if is_aggregated_col:
                    aggregated_col_indices.append(c_idx)
            
            logger.debug(f"Found {len(aggregated_row_indices)} aggregated rows and {len(aggregated_col_indices)} aggregated columns")
            
            # Apply bold font to cells in aggregated rows
            for r_idx in aggregated_row_indices:
                # Bold the row headers for this aggregated row
                for c_idx in range(index_col_start, data_col_start):
                    try:
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.font = AGGREGATED_FONT
                    except IndexError:
                        logger.warning(f"IndexError accessing header cell at row {r_idx}, col {c_idx} during aggregated row styling.")
                        continue
                        
                # Bold the data cells for this aggregated row
                for c_idx in range(data_col_start, max_col + 1):
                    try:
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.font = AGGREGATED_FONT
                    except IndexError:
                        logger.warning(f"IndexError accessing cell at row {r_idx}, col {c_idx} during aggregated row styling.")
                        continue
            
            # Apply bold font to cells in aggregated columns
            for c_idx in aggregated_col_indices:
                # Bold the column headers for this aggregated column
                for r_idx in range(header_row_start, data_row_start):
                    try:
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.font = AGGREGATED_FONT
                    except IndexError:
                        logger.warning(f"IndexError accessing header cell at row {r_idx}, col {c_idx} during aggregated column styling.")
                        continue
                        
                # Bold the data cells for this aggregated column (skipping cells in aggregated rows as they're already bold)
                for r_idx in range(data_row_start, max_row + 1):
                    # Skip rows already processed (aggregated rows)
                    if r_idx in aggregated_row_indices:
                        continue
                    try:
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.font = AGGREGATED_FONT
                    except IndexError:
                        logger.warning(f"IndexError accessing cell at row {r_idx}, col {c_idx} during aggregated column styling.")
                        continue


        logger.info(f"Styling applied successfully to sheet '{worksheet.title}'.")

    except Exception as e:
        logger.error(f"An unexpected error occurred during styling application for sheet '{worksheet.title}': {e}", exc_info=True)
        # Avoid raising StylingError if it's not defined or imported correctly, fallback to generic Exception
        try:
            raise StylingError(f"Failed to apply styles to sheet '{worksheet.title}'. Error: {e}") from e
        except NameError:
             raise Exception(f"Failed to apply styles to sheet '{worksheet.title}'. Error: {e} (StylingError not defined)") from e