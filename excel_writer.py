"""Handles writing DataFrames to Excel files, including the unmerge hack."""

import logging
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet # Explicit import for type hinting
from openpyxl.utils.exceptions import IllegalCharacterError
from typing import Tuple, Optional # Added Optional
from openpyxl.utils.dataframe import dataframe_to_rows

# Assuming ConfigModel and exceptions are importable from sibling modules
from .config import ConfigModel
from .exceptions import ExcelWriterError
from .utils import sanitize_sheet_name # For final sheet name check

# Get the logger instance configured in utils
logger = logging.getLogger("sdc_checker.excel_writer")

# --- Unmerge Hack Function ---
def unmerge_multiindex_table(table: pd.DataFrame) -> pd.DataFrame:
    """
    Unmerge multiindex pandas tables by inserting temporary unique top-level headers.
    Returns the modified table. The calling function is responsible for removing
    these temporary headers from the openpyxl object after writing.

    Args:
        table: The input DataFrame, potentially with MultiIndex rows/columns.

    Returns:
        A modified copy of the DataFrame with temporary unique identifiers added
        as the outermost level for MultiIndex rows and/or columns.
    """
    table_copy = table.copy()
    # Check if columns have more than one level
    if table_copy.columns.nlevels > 1:
        logger.debug("Applying unmerge hack: Adding temporary column level.")
        new_column_levels = [range(table_copy.shape[1])] + [table_copy.columns.get_level_values(i) for i in range(table_copy.columns.nlevels)]
        temp_cols = pd.MultiIndex.from_arrays(new_column_levels)
        temp_cols.names = [None] + table_copy.columns.names
        table_copy.columns = temp_cols
    else:
        logger.debug("Unmerge hack: Columns are single-level, no temporary level added.")

    # Check if index (rows) have more than one level
    if table_copy.index.nlevels > 1:
        logger.debug("Applying unmerge hack: Adding temporary row index level.")
        new_index_levels = [range(table_copy.shape[0])] + [table_copy.index.get_level_values(i) for i in range(table_copy.index.nlevels)]
        temp_rows = pd.MultiIndex.from_arrays(new_index_levels)
        temp_rows.names = [None] + table_copy.index.names
        table_copy.index = temp_rows
    else:
        logger.debug("Unmerge hack: Index is single-level, no temporary level added.")

    return table_copy

# --- Helper Functions ---

def _prepare_workbook_and_sheet(
    output_path: str,
    sheet_name: str,
    existing_workbook: Optional[Workbook] = None
) -> Tuple[Workbook, Worksheet]:
    """
    Uses an existing workbook or loads/creates one, removes the target sheet if it exists,
    and creates/returns the new sheet.

    Args:
        output_path: Path to the Excel file (used only if existing_workbook is None).
        sheet_name: Name of the sheet to prepare.
        existing_workbook: An optional existing openpyxl Workbook object to use.

    Returns:
        A tuple containing the Workbook and the newly created Worksheet.

    Raises:
        ExcelWriterError: If loading/modifying the workbook fails.
    """
    workbook = existing_workbook
    worksheet = None

    if workbook:
        logger.debug(f"Using provided existing workbook object.")
    elif os.path.exists(output_path):
        try:
            logger.debug(f"Loading existing workbook from '{output_path}'.")
            workbook = openpyxl.load_workbook(output_path)
        except Exception as e:
            logger.error(f"Failed to load existing workbook '{output_path}': {e}", exc_info=True)
            raise ExcelWriterError(f"Failed to load existing workbook '{output_path}'. Error: {e}") from e
    else:
        logger.debug(f"Creating new workbook for '{output_path}'.")
        workbook = openpyxl.Workbook()
        # Remove default sheet if we are creating our named sheet as the first one
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
             default_sheet = workbook["Sheet"]
             workbook.remove(default_sheet)
             logger.debug("Removed default 'Sheet'.")

    # Ensure workbook is valid before proceeding
    if not isinstance(workbook, Workbook):
         logger.error(f"Could not obtain a valid Workbook object.")
         raise ExcelWriterError("Internal error: Failed to obtain a valid Workbook object.")

    # Remove existing sheet if it exists in the workbook
    if sheet_name in workbook.sheetnames:
        logger.debug(f"Removing existing sheet '{sheet_name}' from workbook.")
        try:
            sheet_to_remove = workbook[sheet_name]
            workbook.remove(sheet_to_remove)
        except Exception as e:
             logger.error(f"Failed to remove existing sheet '{sheet_name}': {e}", exc_info=True)
             raise ExcelWriterError(f"Failed to remove existing sheet '{sheet_name}'. Error: {e}") from e

    # Create the new sheet
    try:
        logger.debug(f"Creating new sheet '{sheet_name}' in workbook.")
        worksheet = workbook.create_sheet(sheet_name)
    except Exception as e:
        logger.error(f"Failed to create new sheet '{sheet_name}': {e}", exc_info=True)
        raise ExcelWriterError(f"Failed to create new sheet '{sheet_name}'. Error: {e}") from e


    if worksheet is None: # Safeguard
         logger.error(f"Worksheet object for '{sheet_name}' could not be created or assigned.")
         raise ExcelWriterError(f"Internal error: Failed to obtain worksheet object for sheet '{sheet_name}'.")

    return workbook, worksheet

def _write_title_row(worksheet: Worksheet, title_text: str):
    """Writes the title text to cell A1 of the worksheet."""
    title_start_col = 1
    title_cell = worksheet.cell(row=1, column=title_start_col)
    title_cell.value = title_text
    logger.debug(f"Wrote title '{title_text}' to cell {title_cell.coordinate}")

def _write_data_rows(worksheet: Worksheet, df_to_write: pd.DataFrame, temp_col_added: bool, temp_row_added: bool):
    """
    Writes the DataFrame data (including headers and index) to the worksheet,
    skipping temporary rows/columns added by the unmerge hack.
    Data writing starts from Excel row 2.

    Args:
        worksheet: The openpyxl Worksheet to write to.
        df_to_write: The DataFrame (potentially modified by unmerge hack) to write.
        temp_col_added: Flag indicating if a temporary header row was added.
        temp_row_added: Flag indicating if a temporary index column was added.

    Raises:
        ExcelWriterError: If writing data fails (e.g., illegal characters).
    """
    logger.debug(f"Converting DataFrame (shape {df_to_write.shape}) to rows for writing.")
    rows = list(dataframe_to_rows(df_to_write, index=True, header=True))
    logger.debug(f"Using unmerge hack flags: temp_col_added={temp_col_added}, temp_row_added={temp_row_added}")

    logger.debug(f"Writing {len(rows)} potential rows to sheet '{worksheet.title}'...")
    rows_written = 0
    MAX_DEBUG_ROWS = 5 # Log details only for the first few rows

    logger.debug(f"Starting _write_data_rows: total rows to write = {len(rows)}")
    for r_idx, row_data in enumerate(rows):
        logger.debug(f"Row {r_idx}: {row_data}")
        # Skip the first row if it's the temporary header from column unmerge hack
        if r_idx == 0 and temp_col_added:
            logger.debug(f"[Row {r_idx}] Skipping temporary header row (temp_col_added={temp_col_added})")
            continue

        # Calculate target row in Excel (1-based), starting data from row 2
        excel_row = r_idx + 2 - (1 if temp_col_added else 0)
        logger.debug(f"[Row {r_idx}] Writing to Excel row {excel_row}")

        col_cells_written = 0
        for c_idx, value in enumerate(row_data):
            # Skip the first column if it's the temporary index from row unmerge hack
            if c_idx == 0 and temp_row_added:
                logger.debug(f"[Row {r_idx}, Col {c_idx}] Skipping temporary index column (temp_row_added={temp_row_added})")
                continue

            # Calculate target column in Excel (1-based)
            excel_col = c_idx + 1 - (1 if temp_row_added else 0)

            try:
                # Debug: log the type and content of the value before writing
                logger.debug(f"[Row {r_idx}, Col {c_idx}] Original value type: {type(value)}, value: {repr(value)}")
                # If the value is a tuple, log its length and contents
                if isinstance(value, tuple):
                    logger.debug(f"[Row {r_idx}, Col {c_idx}] Tuple detected with length {len(value)}: {value}")
                    # Unwrap 1-element tuples to their scalar content
                    if len(value) == 1:
                        value = value[0]
                        logger.debug(f"[Row {r_idx}, Col {c_idx}] Unwrapped 1-element tuple to scalar: {repr(value)}")
                worksheet.cell(row=excel_row, column=excel_col, value=value)
                safe_value_repr = repr(value)[:50] + ('...' if len(repr(value)) > 50 else '')
                logger.debug(f"[Row {r_idx}, Col {c_idx}] Writing value {safe_value_repr} to Excel cell ({excel_row}, {excel_col})")
                col_cells_written += 1
            except IllegalCharacterError as e:
                safe_value = str(value)[:50] + '...' if len(str(value)) > 50 else str(value)
                logger.error(f"Illegal character error writing value '{safe_value}' to cell ({excel_row}, {excel_col}): {e}", exc_info=True)
                raise ExcelWriterError(f"Failed to write sheet '{worksheet.title}'. Data contains characters unsupported by Excel in cell ({excel_row}, {excel_col}). Value: '{safe_value}'. Error: {e}") from e
            except Exception as e:
                safe_value = str(value)[:50] + '...' if len(str(value)) > 50 else str(value)
                logger.error(f"Error writing value '{safe_value}' to cell ({excel_row}, {excel_col}): {e}", exc_info=True)
                raise ExcelWriterError(f"Failed to write value to cell ({excel_row}, {excel_col}) in sheet '{worksheet.title}'. Value: '{safe_value}'. Error: {e}") from e

        logger.debug(f"[Row {r_idx}] Columns written: {col_cells_written}")
        if col_cells_written > 0:
            rows_written += 1
    logger.debug(f"Finished writing content for {rows_written} data rows to sheet '{worksheet.title}'.")


# --- Main Excel Writing Function ---

def write_excel_sheet(
    df: pd.DataFrame,
    config: ConfigModel,
    title_text: str, # Added parameter
    sheet_name_suffix: str = "", # Existing parameter (kept)
    existing_workbook: Optional[Workbook] = None
) -> Tuple[Workbook, Worksheet]:
    """
    Writes a DataFrame to a specified sheet in an Excel workbook using helper functions.
    Uses an existing workbook object if provided, otherwise loads/creates one.

    Handles:
    - Path/directory preparation.
    - Workbook/sheet preparation via _prepare_workbook_and_sheet.
    - Applying the unmerge hack for multi-index DataFrames.
    - Writing title via _write_title_row.
    - Writing data via _write_data_rows.
    - NOTE: Does NOT save the workbook; saving is handled by the caller.

    Args:
        df: The DataFrame to write.
        config: The validated configuration object.
        title_text: The pre-formatted title string for the sheet.
        sheet_name_suffix: Optional suffix for the sheet name (used for naming only).
        existing_workbook: Optional existing Workbook object to write the sheet into.

    Returns:
        A tuple containing the (potentially modified) Workbook and the Worksheet where data was written.

    Raises:
        ExcelWriterError: If any error occurs during the process.
    """
    output_path = config.output
    base_sheet_name = config.sheet_name

    # Ensure we have a valid base sheet name
    if base_sheet_name is None or base_sheet_name == "":
        base_sheet_name = "Output"  # Fallback default

    proposed_name = base_sheet_name + sheet_name_suffix

    # Final sanitization and truncation check
    final_sheet_name = sanitize_sheet_name(proposed_name)
    if final_sheet_name != proposed_name:
        logger.warning(f"Final sheet name '{proposed_name}' was sanitized/truncated to '{final_sheet_name}'.")

    logger.info(f"Preparing to write sheet '{final_sheet_name}' to file '{output_path}'...")

    workbook = None # Initialize workbook variable

    try:
        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            logger.debug(f"Ensured output directory exists: '{output_dir}'")

        # Apply unmerge hack and determine if temp rows/cols were added
        orig_col_nlevels = df.columns.nlevels
        orig_idx_nlevels = df.index.nlevels
        df_to_write = unmerge_multiindex_table(df)
        temp_col_added = df_to_write.columns.nlevels > orig_col_nlevels
        temp_row_added = df_to_write.index.nlevels > orig_idx_nlevels

        # Prepare workbook and worksheet using helper, passing existing workbook if provided
        workbook, worksheet = _prepare_workbook_and_sheet(
            output_path, final_sheet_name, existing_workbook=existing_workbook
        )

        # Write the provided title text using helper
        _write_title_row(worksheet, title_text)

        # Write data rows using helper
        _write_data_rows(worksheet, df_to_write, temp_col_added, temp_row_added)

        # Workbook saving is now handled by the calling function (__init__.py)
        # after all sheets and styles have been applied.
        logger.info(f"Sheet '{final_sheet_name}' successfully written to workbook object.")
        return workbook, worksheet

    except Exception as e:
        logger.error(f"An unexpected error occurred in write_excel_sheet: {e}", exc_info=True)
        if not isinstance(e, ExcelWriterError):
            raise ExcelWriterError(f"Failed to write Excel sheet '{final_sheet_name}'. Error: {e}") from e
        else:
            raise